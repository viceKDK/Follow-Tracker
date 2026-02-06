import csv
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def popup(title, message):
    """Muestra una ventana emergente con un mensaje."""
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, message)
    root.destroy()

def seleccionar_archivo(titulo):
    """Abre un diálogo para seleccionar un archivo CSV."""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title=titulo,
        filetypes=[("Archivos CSV", "*.csv"), ("Todos los archivos", "*.*")]
    )
    root.destroy()
    return file_path

def leer_csv_usuarios(ruta_csv):
    """Lee un CSV generado por smart_scraper.js y retorna un set de usuarios."""
    usuarios = set()
    try:
        with open(ruta_csv, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            # Intentar saltar encabezado si existe
            header = next(reader, None)
            if header and "Usuario" not in header[0]:
                # Si la primera fila no parece encabezado, reiniciamos (aunque el script JS siempre pone header)
                f.seek(0)
            
            for row in reader:
                if row and len(row) > 0:
                    usuario = row[0].strip()
                    if usuario:
                        usuarios.add(usuario)
    except Exception as e:
        raise Exception(f"Error leyendo {ruta_csv}: {str(e)}")
    return usuarios

def generar_excel(followers, following, output_path):
    """Genera el archivo Excel con el formato específico."""
    
    # Lógica de conjuntos
    nos_seguimos = sorted(followers & following)
    no_me_sigue = sorted(following - followers) # Yo sigo pero él no
    no_lo_sigo = sorted(followers - following)  # Me sigue pero yo no

    # Para esta versión "externa", no tenemos historial, así que "Nuevos" estarán vacíos
    nuevos_seguidores = []
    nuevos_siguiendo = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Seguimiento Instagram"
    ws.sheet_view.showGridLines = False

    # --- 1. TITULO PRINCIPAL (Centrado en B1:H3) ---
    ws.merge_cells("B1:H3")
    ws["B1"] = "Análisis de Perfil Externo"
    ws["B1"].font = Font(size=24, bold=True, color="2E75B6")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    # --- 2. ENCABEZADOS DE TABLAS (Fila 6) ---
    headers = [
        f"Nos seguimos ({len(nos_seguimos)})",
        "",
        f"No lo sigo ({len(no_lo_sigo)})",
        "",
        f"No me sigue ({len(no_me_sigue)})",
        "",
        f"Nuevos Seguidores ({len(nuevos_seguidores)})",
        "",
        f"Nuevos Siguiendo ({len(nuevos_siguiendo)})"
    ]

    for col_idx, text in enumerate(headers, 2): # Columna B es index 2
        cell = ws.cell(row=6, column=col_idx, value=text)
        if text:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            if col_idx == 2: cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Azul
            if col_idx == 4: cell.fill = PatternFill(start_color="E46C0A", end_color="E46C0A", fill_type="solid") # Naranja
            if col_idx == 6: cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Rojo
            if col_idx == 8: cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid") # Verde claro
            if col_idx == 10: cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid") # Celeste

    # --- 3. DATOS (Desde Fila 7) ---
    max_len = max(len(nos_seguimos), len(no_me_sigue), len(no_lo_sigo), 1) # Asegurar al menos 1 fila para tabla válida
    
    for i in range(max_len):
        row_data = [
            "", # Columna A
            nos_seguimos[i] if i < len(nos_seguimos) else "",
            "",
            no_lo_sigo[i] if i < len(no_lo_sigo) else "",
            "",
            no_me_sigue[i] if i < len(no_me_sigue) else "",
            "",
            "", # Nuevos
            "",
            ""  # Nuevos
        ]
        ws.append(row_data)

    # --- 4. FORMATO ---
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 30

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Tablas
    table_range_b = f"B6:B{max(7, max_len + 6)}"
    table_range_d = f"D6:D{max(7, max_len + 6)}"
    table_range_f = f"F6:F{max(7, max_len + 6)}"
    # Rangos vacíos para nuevos (Excel requiere que la tabla tenga datos o encabezado, si está vacía puede dar error,
    # pero openpyxl suele manejarlo bien si definimos el rango. Si están vacíos, mejor no crear tabla o crearla con filas vacías)
    
    # Solo creamos tablas para las 3 primeras columnas que tienen datos seguros
    tables_config = [
        ("Tab_NosSeg", table_range_b, "TableStyleMedium2"),
        ("Tab_NoLoSig", table_range_d, "TableStyleMedium3"),
        ("Tab_NoMeSig", table_range_f, "TableStyleMedium7")
    ]

    for name, ref, style in tables_config:
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
        ws.add_table(tab)
        
        # Bordes
        col_letter = ref[0]
        start_row = 6
        end_row = max(7, max_len + 6)
        for row in range(start_row, end_row + 1):
            ws[f"{col_letter}{row}"].border = thin_border

    wb.save(output_path)
    return len(nos_seguimos), len(no_lo_sigo), len(no_me_sigue)

def main():
    try:
        # 1. Seleccionar CSV de Seguidores
        path_followers = seleccionar_archivo("Selecciona el CSV de SEGUIDORES (Followers)")
        if not path_followers:
            return # Usuario canceló

        # 2. Seleccionar CSV de Siguiendo
        path_following = seleccionar_archivo("Selecciona el CSV de SIGUIENDO (Following)")
        if not path_following:
            return # Usuario canceló

        print("Procesando archivos...")
        
        followers = leer_csv_usuarios(path_followers)
        following = leer_csv_usuarios(path_following)

        # Determinar nombre de salida
        dir_path = os.path.dirname(path_followers)
        output_path = os.path.join(dir_path, "analisis_externo.xlsx")

        c_nos, c_nolosigo, c_nomeshigue = generar_excel(followers, following, output_path)

        popup("Éxito", f"Análisis completado.\n\nArchivo guardado en:\n{output_path}\n\nResumen:\n- Nos seguimos: {c_nos}\n- No lo sigo: {c_nolosigo}\n- No me sigue: {c_nomeshigue}")

    except Exception as e:
        popup("Error", f"Ocurrió un error:\n{str(e)}")

if __name__ == "__main__":
    main()
