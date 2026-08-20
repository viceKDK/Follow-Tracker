import csv
import glob
import hashlib
import json
import os
import re
import shutil
import sys
import tkinter as tk
from datetime import datetime
import time
import tempfile
from tkinter import filedialog, messagebox, simpledialog

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


AUTO_CSV_RE = re.compile(
    r"^ig_auto_(.+)_(followers|following)_([a-z0-9-]+)_(\d+)\.csv$",
    re.IGNORECASE,
)
LEGACY_AUTO_CSV_RE = re.compile(
    r"^ig_auto_(.+)_(followers|following)_(\d+)\.csv$",
    re.IGNORECASE,
)
LEGACY_PAIR_MAX_GAP_MS = 15 * 60 * 1000


def _with_hidden_root():
    root = tk.Tk()
    root.withdraw()
    return root


def popup(title, message):
    root = _with_hidden_root()
    messagebox.showinfo(title, message, parent=root)
    root.destroy()


def confirmar_actualizacion_historial(message):
    root = _with_hidden_root()
    confirmed = messagebox.askyesno(
        "Cambio grande detectado",
        f"{message}\n\nEl reporte se genero igualmente. "
        "¿Quieres usar esta captura como nueva linea base?",
        parent=root,
    )
    root.destroy()
    return confirmed


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)
    return path


def format_dt(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def format_unix_ms(ts_ms):
    try:
        dt = datetime.fromtimestamp(ts_ms / 1000.0)
        return format_dt(dt)
    except Exception:
        return format_dt(datetime.now())


def sanitize_folder_name(text):
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", str(text or "")).strip("._")
    return safe or "perfil"


def seleccionar_archivo(titulo):
    root = _with_hidden_root()
    file_path = filedialog.askopenfilename(
        title=titulo,
        filetypes=[("JSON or CSV", "*.json *.csv"), ("All files", "*.*")],
        parent=root,
    )
    root.destroy()
    return file_path


def preguntar_perfil(valor_inicial=""):
    root = _with_hidden_root()
    value = simpledialog.askstring(
        "Perfil de Instagram",
        "Escribe el usuario del perfil para mantener un historial correcto:",
        initialvalue=valor_inicial,
        parent=root,
    )
    root.destroy()
    if value is None:
        return None
    value = value.strip().lstrip("@").lower()
    if not re.fullmatch(r"[a-zA-Z0-9._]+", value):
        raise ValueError("El usuario del perfil solo puede contener letras, numeros, puntos y guiones bajos.")
    return value


def preguntar_modo():
    selected_mode = {"value": None}

    root = tk.Tk()
    root.title("Follow Tracker - Inicio")
    root.geometry("1000x720")
    root.minsize(900, 640)
    try:
        root.state("zoomed")
    except Exception:
        pass
    root.configure(bg="#f4f7fb")

    card = tk.Frame(root, bg="white", bd=1, relief="solid")
    card.pack(fill="both", expand=True, padx=16, pady=16)

    title = tk.Label(
        card,
        text="Elige como quieres analizar Instagram",
        font=("Segoe UI", 15, "bold"),
        bg="white",
        fg="#1d3557",
    )
    title.pack(anchor="w", padx=18, pady=(16, 8))

    subtitle = tk.Label(
        card,
        text="El flujo recomendado para cuentas externas es AUTO 1-click (esperar).",
        font=("Segoe UI", 10),
        bg="white",
        fg="#4a5568",
        wraplength=700,
        justify="left",
    )
    subtitle.pack(anchor="w", padx=18, pady=(0, 12))

    my_account_box = tk.LabelFrame(
        card,
        text="Modo 1: Mi cuenta (JSON)",
        font=("Segoe UI", 10, "bold"),
        bg="white",
        fg="#1f7a8c",
        padx=12,
        pady=10,
    )
    my_account_box.pack(fill="x", padx=18, pady=(0, 10))

    my_account_text = (
        "Requisitos:\n"
        "- Crear carpeta yo dentro del proyecto.\n"
        "- Poner ahi: followers_1.json y following.json.\n"
        "- Se genera yo/seguidores_vs_seguidos.xlsx."
    )
    tk.Label(
        my_account_box,
        text=my_account_text,
        font=("Segoe UI", 10),
        bg="white",
        fg="#2d3748",
        justify="left",
        anchor="w",
    ).pack(fill="x")

    auto_box = tk.LabelFrame(
        card,
        text="Modo 2: Otra cuenta (AUTO 1-click)",
        font=("Segoe UI", 10, "bold"),
        bg="white",
        fg="#155724",
        padx=12,
        pady=10,
    )
    auto_box.pack(fill="x", padx=18, pady=(0, 10))

    auto_text = (
        "Flujo automatico recomendado:\n"
        "1. Deja abierto el perfil de la persona en Instagram.\n"
        "2. Ejecuta smart_scraper.js (hace seguidores y luego seguidos solo).\n"
        "3. Abre esta app y pulsa AUTO.\n\n"
        "La app detecta el ultimo par de CSV, crea carpeta por usuario,\n"
        "copia ahi ambos CSV y genera el Excel automaticamente."
    )
    tk.Label(
        auto_box,
        text=auto_text,
        font=("Segoe UI", 10),
        bg="white",
        fg="#2d3748",
        justify="left",
        anchor="w",
        wraplength=740,
    ).pack(fill="x")

    buttons = tk.Frame(card, bg="white")
    buttons.pack(fill="x", padx=18, pady=(10, 16))

    def choose_local():
        selected_mode["value"] = "local"
        root.destroy()

    def choose_auto_wait():
        selected_mode["value"] = "auto_wait"
        root.destroy()

    def close_app():
        selected_mode["value"] = None
        root.destroy()

    tk.Button(
        buttons,
        text="Usar mi cuenta (JSON)",
        font=("Segoe UI", 10, "bold"),
        bg="#1f7a8c",
        fg="white",
        activebackground="#16697a",
        activeforeground="white",
        padx=14,
        pady=8,
        relief="flat",
        command=choose_local,
        cursor="hand2",
    ).pack(side="left", padx=(0, 10))

    tk.Button(
        buttons,
        text="AUTO 1-click (esperar)",
        font=("Segoe UI", 10, "bold"),
        bg="#0f766e",
        fg="white",
        activebackground="#115e59",
        activeforeground="white",
        padx=14,
        pady=8,
        relief="flat",
        command=choose_auto_wait,
        cursor="hand2",
    ).pack(side="left")

    tk.Button(
        buttons,
        text="Salir",
        font=("Segoe UI", 10),
        bg="#e2e8f0",
        fg="#1a202c",
        activebackground="#cbd5e0",
        activeforeground="#1a202c",
        padx=12,
        pady=8,
        relief="flat",
        command=close_app,
        cursor="hand2",
    ).pack(side="right")

    root.protocol("WM_DELETE_WINDOW", close_app)
    root.mainloop()
    return selected_mode["value"]


def cargar_json(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"No se encontro el archivo: {os.path.basename(path)}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def leer_csv_usuarios(path):
    usuarios = set()
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row:
                continue
            usuario = row[0].strip().lstrip("@")
            if not usuario or usuario.lower() in {"usuario", "username"}:
                continue
            # Strictly accept Instagram usernames.  In particular, never pass
            # spreadsheet formula-like values through to openpyxl.
            if usuario[0] in "=+-@" or not re.fullmatch(r"[A-Za-z0-9._]+", usuario):
                raise ValueError(f"Username invalido en {os.path.basename(path)}: {usuario!r}")
            usuarios.add(usuario.lower())
    if not usuarios:
        raise ValueError(f"CSV vacio o sin usernames validos: {path}")
    return usuarios


def parse_followers_json(path):
    raw = cargar_json(path)
    values = {
        str(item["string_list_data"][0]["value"]).strip().lstrip("@").lower()
        for item in raw
        if item.get("string_list_data")
        and item["string_list_data"]
        and item["string_list_data"][0].get("value")
    }
    return _validate_user_set(values, "followers JSON")


def parse_following_json(path):
    raw = cargar_json(path)
    rel = raw.get("relationships_following", [])
    values = {
        str(item["string_list_data"][0]["value"]).strip().lstrip("@").lower()
        for item in rel
        if item.get("string_list_data")
        and item["string_list_data"]
        and item["string_list_data"][0].get("value")
    }
    return _validate_user_set(values, "following JSON")


def historial_path(base_path, key):
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:10]
    return os.path.join(base_path, f"historial_{digest}.json")


def cargar_historial(path):
    if not os.path.exists(path):
        return set(), set()
    return cargar_historial_estricto(path)


def _validate_user_set(values, label="usuarios"):
    if not isinstance(values, (set, list, tuple)):
        raise ValueError(f"{label} invalido")
    result = set()
    for value in values:
        value = str(value).strip().lstrip("@").lower()
        if not re.fullmatch(r"[a-z0-9._]+", value):
            raise ValueError(f"{label} contiene username invalido")
        result.add(value)
    return result


def historial_estado(path):
    """Return missing, valid, or corrupt without silently treating corruption as baseline."""
    if not os.path.isfile(path):
        return "missing"
    try:
        cargar_historial_estricto(path)
    except Exception:
        return "corrupt"
    return "valid"


def cargar_historial_estricto(path):
    data = cargar_json(path)
    if data.get("schema_version") != 2:
        raise ValueError("schema de historial invalido")
    return (_validate_user_set(data["followers"]), _validate_user_set(data["following"]))


def historial_existe(path):
    return os.path.isfile(path)


def calcular_cambios(actual, anterior, hay_historial):
    if not hay_historial:
        return [], []
    return sorted(actual - anterior), sorted(anterior - actual)


def guardar_historial(path, followers, following):
    ensure_dir(os.path.dirname(path) or ".")
    fd, tmp_path = tempfile.mkstemp(prefix=".historial-", suffix=".tmp", dir=os.path.dirname(path) or ".")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "schema_version": 2,
                    "followers": sorted(followers),
                    "following": sorted(following),
                    "updated_at": format_dt(datetime.now()),
                },
                f,
                ensure_ascii=False,
                indent=2,
            )
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def validar_snapshot(followers, following, previous_followers=None, previous_following=None):
    followers = _validate_user_set(followers, "followers")
    following = _validate_user_set(following, "following")
    if previous_followers is not None:
        for current, previous, label in ((followers, previous_followers, "followers"), (following, previous_following, "following")):
            if previous and not current:
                raise ValueError(f"Caida sospechosa: {label} quedo vacio")
            if len(previous) >= 10 and len(current) < len(previous) * 0.5:
                raise ValueError(f"Caida sospechosa de {label}: {len(previous)} -> {len(current)}")
    return followers, following


def detectar_caida_sospechosa(followers, following, previous_followers, previous_following):
    warnings = []
    for current, previous, label in (
        (followers, previous_followers, "seguidores"),
        (following, previous_following, "seguidos"),
    ):
        if previous and not current:
            warnings.append(f"{label}: {len(previous)} -> 0")
        elif len(previous) >= 10 and len(current) < len(previous) * 0.5:
            warnings.append(f"{label}: {len(previous)} -> {len(current)}")
    if not warnings:
        return None
    return "La captura presenta una caida mayor al 50% (" + ", ".join(warnings) + ")."


def generar_excel(
    output_path,
    titulo,
    followers,
    following,
    nuevos_seguidores,
    nuevos_siguiendo,
    ultimo_scrapeo,
    dejaron_de_seguir=None,
    dejaste_de_seguir=None,
):
    dejaron_de_seguir = sorted(dejaron_de_seguir or [])
    dejaste_de_seguir = sorted(dejaste_de_seguir or [])
    nos_seguimos = sorted(followers & following)
    no_me_sigue = sorted(following - followers)
    no_lo_sigo = sorted(followers - following)

    wb = Workbook()
    ws = wb.active
    ws.title = "Seguimiento Instagram"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:P3")
    ws["B1"] = titulo
    ws["B1"].font = Font(size=24, bold=True, color="2E75B6")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        f"Nos seguimos ({len(nos_seguimos)})",
        "",
        f"No lo sigo ({len(no_lo_sigo)})",
        "",
        f"No me sigue ({len(no_me_sigue)})",
        "",
        f"Nuevos Seguidores ({len(nuevos_seguidores)})",
        "",
        f"Nuevos Siguiendo ({len(nuevos_siguiendo)})",
        "",
        f"Dejaron de Seguirte ({len(dejaron_de_seguir)})",
        "",
        f"Dejaste de Seguir ({len(dejaste_de_seguir)})",
        "",
        "Ultimo Scrapeo",
    ]

    for col_idx, text in enumerate(headers, 2):
        cell = ws.cell(row=6, column=col_idx, value=text)
        if not text:
            continue
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        if col_idx == 2:
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        if col_idx == 4:
            cell.fill = PatternFill(start_color="E46C0A", end_color="E46C0A", fill_type="solid")
        if col_idx == 6:
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        if col_idx == 8:
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        if col_idx == 10:
            cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        if col_idx == 12:
            cell.fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
        if col_idx == 14:
            cell.fill = PatternFill(start_color="9C2F2F", end_color="9C2F2F", fill_type="solid")
        if col_idx == 16:
            cell.fill = PatternFill(start_color="795548", end_color="795548", fill_type="solid")

    max_len = max(
        len(nos_seguimos),
        len(no_me_sigue),
        len(no_lo_sigo),
        len(nuevos_seguidores),
        len(nuevos_siguiendo),
        len(dejaron_de_seguir),
        len(dejaste_de_seguir),
        1,
    )
    for i in range(max_len):
        row_data = [
            "",
            nos_seguimos[i] if i < len(nos_seguimos) else "",
            "",
            no_lo_sigo[i] if i < len(no_lo_sigo) else "",
            "",
            no_me_sigue[i] if i < len(no_me_sigue) else "",
            "",
            nuevos_seguidores[i] if i < len(nuevos_seguidores) else "",
            "",
            nuevos_siguiendo[i] if i < len(nuevos_siguiendo) else "",
            "",
            dejaron_de_seguir[i] if i < len(dejaron_de_seguir) else "",
            "",
            dejaste_de_seguir[i] if i < len(dejaste_de_seguir) else "",
            "",
            ultimo_scrapeo if i == 0 else "",
        ]
        ws.append(row_data)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 30
    ws.column_dimensions["M"].width = 10
    ws.column_dimensions["N"].width = 30
    ws.column_dimensions["O"].width = 10
    ws.column_dimensions["P"].width = 24

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    table_defs = [
        ("Tab_NosSeg", f"B6:B{max(7, max_len + 6)}", "TableStyleMedium2"),
        ("Tab_NoLoSig", f"D6:D{max(7, max_len + 6)}", "TableStyleMedium3"),
        ("Tab_NoMeSig", f"F6:F{max(7, max_len + 6)}", "TableStyleMedium7"),
        ("Tab_NuevosSeg", f"H6:H{max(7, max_len + 6)}", "TableStyleMedium9"),
        ("Tab_NuevosSig", f"J6:J{max(7, max_len + 6)}", "TableStyleMedium11"),
        ("Tab_DejaronSeg", f"L6:L{max(7, max_len + 6)}", "TableStyleMedium4"),
        ("Tab_DejasteSeg", f"N6:N{max(7, max_len + 6)}", "TableStyleMedium5"),
        ("Tab_UltimoScrapeo", "P6:P7", "TableStyleMedium6"),
    ]
    for name, ref, style in table_defs:
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
        ws.add_table(tab)
        col_letter = ref[0]
        end_row = int(ref.split(":")[1][1:])
        for row in range(6, end_row + 1):
            ws[f"{col_letter}{row}"].border = thin_border

    wb.save(output_path)
    return len(nos_seguimos), len(no_lo_sigo), len(no_me_sigue)


def cargar_modo_mi_cuenta(base_path):
    yo_dir = ensure_dir(os.path.join(base_path, "yo"))
    followers_paths = sorted(glob.glob(os.path.join(yo_dir, "followers*.json")))
    following_paths = sorted(glob.glob(os.path.join(yo_dir, "following*.json")))
    if not followers_paths or not following_paths:
        raise FileNotFoundError("Se necesitan followers*.json y following*.json en la carpeta yo")
    followers = set().union(*(parse_followers_json(p) for p in followers_paths))
    following = set().union(*(parse_following_json(p) for p in following_paths))
    validar_snapshot(followers, following)

    profile = preguntar_perfil("")
    if profile is None:
        return None
    h_path = historial_path(base_path, f"local::{profile}")
    state = historial_estado(h_path)
    if state == "corrupt":
        raise ValueError("El historial local esta corrupto; no se reemplazara automaticamente.")
    has_history = state == "valid"
    followers_prev, following_prev = cargar_historial_estricto(h_path) if has_history else (set(), set())
    history_warning = detectar_caida_sospechosa(
        followers, following, followers_prev, following_prev
    ) if has_history else None

    nuevos_seguidores, dejaron_de_seguir = calcular_cambios(followers, followers_prev, has_history)
    nuevos_siguiendo, dejaste_de_seguir = calcular_cambios(following, following_prev, has_history)

    return {
        "titulo": "Seguimiento de Instagram",
        "output_path": os.path.join(yo_dir, "seguidores_vs_seguidos.xlsx"),
        "work_dir": yo_dir,
        "followers": followers,
        "following": following,
        "nuevos_seguidores": nuevos_seguidores,
        "nuevos_siguiendo": nuevos_siguiendo,
        "dejaron_de_seguir": dejaron_de_seguir,
        "dejaste_de_seguir": dejaste_de_seguir,
        "has_history": has_history,
        "modo_label": "mi cuenta (JSON)",
        "ultimo_scrapeo": format_dt(datetime.now()),
        "history_path": h_path,
        "history_warning": history_warning,
    }


def cargar_modo_otra_cuenta(base_path):
    path_followers = seleccionar_archivo("Selecciona CSV de seguidores (cuenta externa)")
    if not path_followers:
        return None
    path_following = seleccionar_archivo("Selecciona CSV de seguidos (cuenta externa)")
    if not path_following:
        return None

    followers = leer_csv_usuarios(path_followers)
    following = leer_csv_usuarios(path_following)

    parsed_followers = _parse_auto_csv_name(path_followers)
    parsed_following = _parse_auto_csv_name(path_following)
    initial_profile = ""
    if parsed_followers and parsed_following and parsed_followers["profile"] == parsed_following["profile"]:
        initial_profile = parsed_followers["profile"]
    profile = preguntar_perfil(initial_profile)
    if profile is None:
        return None

    profile_key = f"external::{profile}"
    h_path = historial_path(base_path, profile_key)
    state = historial_estado(h_path)
    if state == "corrupt":
        raise ValueError("El historial externo esta corrupto; no se reemplazara automaticamente.")
    has_history = state == "valid"
    followers_prev, following_prev = cargar_historial_estricto(h_path) if has_history else (set(), set())
    history_warning = detectar_caida_sospechosa(
        followers, following, followers_prev, following_prev
    ) if has_history else None

    nuevos_seguidores, dejaron_de_seguir = calcular_cambios(followers, followers_prev, has_history)
    nuevos_siguiendo, dejaste_de_seguir = calcular_cambios(following, following_prev, has_history)

    out_dir = os.path.dirname(path_followers) or base_path
    output_path = os.path.join(out_dir, "analisis_externo.xlsx")
    return {
        "titulo": f"Analisis de Perfil Externo ({profile})",
        "output_path": output_path,
        "work_dir": out_dir,
        "followers": followers,
        "following": following,
        "nuevos_seguidores": nuevos_seguidores,
        "nuevos_siguiendo": nuevos_siguiendo,
        "dejaron_de_seguir": dejaron_de_seguir,
        "dejaste_de_seguir": dejaste_de_seguir,
        "has_history": has_history,
        "modo_label": "otra cuenta (CSV)",
        "ultimo_scrapeo": format_dt(datetime.now()),
        "history_path": h_path,
        "history_warning": history_warning,
    }


def _parse_auto_csv_name(path):
    name = os.path.basename(path)
    # Formato actual: ig_auto_<profile>_<phase>_<run_id>_<timestamp>.csv
    m = AUTO_CSV_RE.match(name)
    if m:
        profile, phase, run_id, ts = m.group(1), m.group(2).lower(), m.group(3).lower(), int(m.group(4))
        return {"path": path, "profile": profile, "phase": phase, "run_id": run_id, "ts": ts}
    # Compatibilidad temporal con descargas anteriores. Solo se emparejan si
    # sus timestamps estan suficientemente cerca.
    legacy = LEGACY_AUTO_CSV_RE.match(name)
    if legacy:
        profile, phase, ts = legacy.group(1), legacy.group(2).lower(), int(legacy.group(3))
        return {"path": path, "profile": profile, "phase": phase, "run_id": None, "ts": ts}
    return None


def _collect_auto_csv_candidates(base_path):
    candidates = []
    roots = [base_path, os.path.join(os.path.expanduser("~"), "Downloads")]
    for root in roots:
        if not root or not os.path.isdir(root):
            continue
        pattern = os.path.join(root, "ig_auto_*_*.csv")
        for p in glob.glob(pattern):
            parsed = _parse_auto_csv_name(p)
            if parsed:
                candidates.append(parsed)
    return candidates


def _select_best_auto_pair(candidates):
    complete_pairs = []

    by_run = {}
    for item in candidates:
        if item.get("run_id"):
            by_run.setdefault((item["profile"], item["run_id"]), []).append(item)
    for (profile, run_id), items in by_run.items():
        followers = max((x for x in items if x["phase"] == "followers"), key=lambda x: x["ts"], default=None)
        following = max((x for x in items if x["phase"] == "following"), key=lambda x: x["ts"], default=None)
        if followers and following:
            complete_pairs.append(
                {
                    "profile": profile,
                    "run_id": run_id,
                    "followers_path": followers["path"],
                    "following_path": following["path"],
                    "ts": max(followers["ts"], following["ts"]),
                }
            )

    # Archivos antiguos no tienen run_id. Elegimos el par mas reciente dentro
    # de una ventana corta; nunca mezclamos automaticamente archivos muy alejados.
    legacy_by_profile = {}
    for item in candidates:
        if not item.get("run_id"):
            legacy_by_profile.setdefault(item["profile"], []).append(item)
    for profile, items in legacy_by_profile.items():
        followers_items = [x for x in items if x["phase"] == "followers"]
        following_items = [x for x in items if x["phase"] == "following"]
        possible = [
            (max(f["ts"], g["ts"]), abs(f["ts"] - g["ts"]), f, g)
            for f in followers_items
            for g in following_items
            if abs(f["ts"] - g["ts"]) <= LEGACY_PAIR_MAX_GAP_MS
        ]
        if possible:
            _, _, followers, following = max(possible, key=lambda x: (x[0], -x[1]))
            complete_pairs.append(
                {
                    "profile": profile,
                    "run_id": None,
                    "followers_path": followers["path"],
                    "following_path": following["path"],
                    "ts": max(followers["ts"], following["ts"]),
                }
            )

    return max(complete_pairs, key=lambda x: x["ts"], default=None)


def _find_latest_auto_pair(base_path):
    return _select_best_auto_pair(_collect_auto_csv_candidates(base_path))


def _find_new_auto_pair_since(base_path, start_ts_ms):
    candidates = [x for x in _collect_auto_csv_candidates(base_path) if x["ts"] >= start_ts_ms]
    return _select_best_auto_pair(candidates)


def _build_auto_data_from_pair(base_path, pair):
    followers = leer_csv_usuarios(pair["followers_path"])
    following = leer_csv_usuarios(pair["following_path"])

    validar_snapshot(followers, following)
    profile_folder = ensure_dir(os.path.join(base_path, sanitize_folder_name(pair["profile"])))
    target_followers = os.path.join(profile_folder, "followers.csv")
    target_following = os.path.join(profile_folder, "following.csv")
    shutil.copy2(pair["followers_path"], target_followers)
    shutil.copy2(pair["following_path"], target_following)

    profile_key = f"external_auto::{pair['profile']}"
    h_path = historial_path(base_path, profile_key)
    state = historial_estado(h_path)
    if state == "corrupt":
        raise ValueError("El historial AUTO esta corrupto; no se reemplazara automaticamente.")
    has_history = state == "valid"
    followers_prev, following_prev = cargar_historial_estricto(h_path) if has_history else (set(), set())
    history_warning = detectar_caida_sospechosa(
        followers, following, followers_prev, following_prev
    ) if has_history else None

    nuevos_seguidores, dejaron_de_seguir = calcular_cambios(followers, followers_prev, has_history)
    nuevos_siguiendo, dejaste_de_seguir = calcular_cambios(following, following_prev, has_history)

    output_path = os.path.join(profile_folder, "seguidores_vs_seguidos.xlsx")
    return {
        "titulo": f"Analisis de Perfil Externo ({pair['profile']})",
        "output_path": output_path,
        "work_dir": profile_folder,
        "followers": followers,
        "following": following,
        "nuevos_seguidores": nuevos_seguidores,
        "nuevos_siguiendo": nuevos_siguiendo,
        "dejaron_de_seguir": dejaron_de_seguir,
        "dejaste_de_seguir": dejaste_de_seguir,
        "has_history": has_history,
        "modo_label": "otra cuenta (AUTO)",
        "ultimo_scrapeo": format_unix_ms(pair["ts"]),
        "copied_files": [target_followers, target_following],
        "history_path": h_path,
        "history_warning": history_warning,
    }


def cargar_modo_otra_cuenta_auto(base_path):
    pair = _find_latest_auto_pair(base_path)
    if not pair:
        raise FileNotFoundError(
            "No se encontro un par automatico de CSV.\n"
            "Archivos esperados: ig_auto_<perfil>_followers_<run_id>_<ts>.csv y "
            "ig_auto_<perfil>_following_<run_id>_<ts>.csv\n"
            "en la carpeta del programa o en Downloads."
        )

    return _build_auto_data_from_pair(base_path, pair)


def cargar_modo_otra_cuenta_auto_wait(base_path, timeout_seconds=60 * 45, poll_seconds=2):
    popup(
        "AUTO 1-click",
        "La app esperara a que termines el scraper en Instagram.\n\n"
        "Pasos ahora:\n"
        "1. Deja esta app abierta.\n"
        "2. Ve a Instagram perfil objetivo.\n"
        "3. Ejecuta smart_scraper.js.\n"
        "4. Al terminar, el Excel se generara solo.",
    )
    start_ts_ms = int(time.time() * 1000)
    start = time.time()
    while (time.time() - start) <= timeout_seconds:
        pair = _find_new_auto_pair_since(base_path, start_ts_ms)
        if pair:
            return _build_auto_data_from_pair(base_path, pair)
        time.sleep(poll_seconds)
    raise TimeoutError("No se detectaron 2 CSV nuevos dentro del tiempo limite.")


def procesar_datos():
    try:
        if getattr(sys, "frozen", False):
            base_path = os.path.dirname(sys.executable)
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))

        choice = preguntar_modo()
        if choice is None:
            return

        if choice == "local":
            data = cargar_modo_mi_cuenta(base_path)
        elif choice == "auto_wait":
            data = cargar_modo_otra_cuenta_auto_wait(base_path)
        else:
            data = cargar_modo_otra_cuenta(base_path)
            if data is None:
                return

        if data is None:
            return

        c_nos, c_no_lo_sigo, c_no_me_sigue = generar_excel(
            data["output_path"],
            data["titulo"],
            data["followers"],
            data["following"],
            data["nuevos_seguidores"],
            data["nuevos_siguiendo"],
            data["ultimo_scrapeo"],
            data.get("dejaron_de_seguir", []),
            data.get("dejaste_de_seguir", []),
        )
        # El reporte es el punto de commit. Una caida grande necesita
        # confirmacion explicita antes de reemplazar la linea base.
        history_saved = True
        if data.get("history_warning"):
            history_saved = confirmar_actualizacion_historial(data["history_warning"])
        if history_saved:
            guardar_historial(data["history_path"], data["followers"], data["following"])

        if data.get("has_history"):
            msg_nuevos = (
                f"\n\nComparacion con anterior:\n"
                f"- Nuevos seguidores: {len(data['nuevos_seguidores'])}\n"
                f"- Nuevos siguiendo: {len(data['nuevos_siguiendo'])}\n"
                f"- Dejaron de seguirte: {len(data.get('dejaron_de_seguir', []))}\n"
                f"- Dejaste de seguir: {len(data.get('dejaste_de_seguir', []))}"
            )
            if not history_saved:
                msg_nuevos += "\n- Historial: se conservo la linea base anterior."
        elif history_saved:
            msg_nuevos = "\n\n(Primera ejecucion: se guardo la linea base)"
        else:
            msg_nuevos = "\n\nNo se guardo una linea base."

        files_msg = ""
        if data.get("copied_files"):
            files_msg = (
                "\nCSV guardados en carpeta del usuario:\n"
                f"- {data['copied_files'][0]}\n"
                f"- {data['copied_files'][1]}\n"
            )

        popup(
            "Listo",
            f"Modo: {data['modo_label']}\n"
            f"Carpeta de salida:\n{data.get('work_dir', os.path.dirname(data['output_path']))}\n"
            f"Archivo generado:\n{data['output_path']}\n\n"
            f"Resumen:\n"
            f"- Nos seguimos: {c_nos}\n"
            f"- No lo sigo: {c_no_lo_sigo}\n"
            f"- No me sigue: {c_no_me_sigue}"
            f"{msg_nuevos}\n"
            f"Ultimo scrapeo: {data['ultimo_scrapeo']}"
            f"{files_msg}",
        )

    except FileNotFoundError as e:
        popup("Archivo faltante", str(e))
    except TimeoutError as e:
        popup("Tiempo agotado", str(e))
    except Exception as e:
        popup("Error", f"Ocurrio un error inesperado:\n{str(e)}")


if __name__ == "__main__":
    procesar_datos()
